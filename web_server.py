import json
import asyncio
import uvicorn
import io
from typing import Optional, Dict, Any, List

from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook

from report.dailyreport_llm_mcp import generate_daily_report
from report.heatingreport_llm_mcp import generate_heating_report
from report.emergencyreport_llm_mcp import generate_emergency_report
from services.dify_proxy_service import dify_proxy

app = FastAPI(title="12345 智能报表生成系统")

DAILY_MCP_ENTRY = "http://127.0.0.1:9001/daily_report_mcp"
HEATING_MCP_ENTRY = "http://127.0.0.1:9002/heating_report_mcp"
EMERGENCY_MCP_ENTRY = "http://127.0.0.1:9003/emergency_report_mcp"

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# --------- Requests ---------
class GenerateStreamRequest(BaseModel):
    report_type: str  # "daily" | "heating" | "emergency"
    date: Optional[str] = None  # YYYY-MM-DD for daily/emergency
    year: Optional[int] = None  # year for heating


class ReportRequest(BaseModel):
    date: str  # YYYY-MM-DD


class HeatingReportRequest(BaseModel):
    year: int


# --------- Dify API 代理请求模型 ---------
class DifyWorkflowRequest(BaseModel):
    app_type: str  # order_recognition, element_extraction, dispatch_assistant, address_recognition
    inputs: Dict[str, Any]
    response_mode: str = "blocking"  # blocking 或 streaming
    user: Optional[str] = None
    conversation_id: Optional[str] = None


# --------- Static ---------
@app.get("/")
async def root_index():
    return FileResponse("static/index.html")


app.mount("/static", StaticFiles(directory="static"), name="static")


# --------- SSE helpers ---------
def sse_pack(obj: Dict[str, Any]) -> str:
    return "data: " + json.dumps(obj, ensure_ascii=False) + "\n\n"


@app.post("/api/generate/stream")
async def generate_stream(req: GenerateStreamRequest):
    """
    统一 SSE 接口：
    - daily: {report_type:"daily", date:"YYYY-MM-DD"}
    - emergency: {report_type:"emergency", date:"YYYY-MM-DD"}
    - heating: {report_type:"heating", year:2024}
    """
    q: asyncio.Queue = asyncio.Queue()

    async def event_cb(evt: Dict[str, Any]):
        await q.put({"type": "event", "data": evt})

    async def worker():
        try:
            rt = (req.report_type or "").strip().lower()

            if rt == "daily":
                if not req.date:
                    raise ValueError("daily 需要 date (YYYY-MM-DD)")
                report = await generate_daily_report(req.date, DAILY_MCP_ENTRY, event_cb=event_cb)
                await q.put({"type": "final", "data": {"report": report}})

            elif rt == "emergency":
                if not req.date:
                    raise ValueError("emergency 需要 date (YYYY-MM-DD)")
                report = await generate_emergency_report(req.date, EMERGENCY_MCP_ENTRY, event_cb=event_cb)
                await q.put({"type": "final", "data": {"report": report}})

            elif rt == "heating":
                if req.year is None:
                    raise ValueError("heating 需要 year (int)")
                report = await generate_heating_report(req.year, HEATING_MCP_ENTRY, event_cb=event_cb)
                await q.put({"type": "final", "data": {"report": report}})

            else:
                raise ValueError(f"未知 report_type: {req.report_type}")

        except Exception as e:
            await q.put({"type": "error", "data": {"message": str(e)}})
        finally:
            await q.put(None)

    async def gen():
        task = asyncio.create_task(worker())
        try:
            while True:
                item = await q.get()
                if item is None:
                    break
                yield sse_pack(item)
        finally:
            if not task.done():
                task.cancel()

    return StreamingResponse(gen(), media_type="text/event-stream")


# --------- Non-streaming (保留原接口) ---------
@app.post("/api/generate_report")
async def generate_report_api(request: ReportRequest):
    try:
        report_content = await generate_daily_report(request.date, DAILY_MCP_ENTRY)
        return {"status": "success", "report": report_content}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/generate_heating_report")
async def generate_heating_report_api(request: HeatingReportRequest):
    try:
        report_content = await generate_heating_report(request.year, HEATING_MCP_ENTRY)
        return {"status": "success", "report": report_content}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/generate_emergency_report")
async def generate_emergency_report_api(request: ReportRequest):
    try:
        report_content = await generate_emergency_report(request.date, EMERGENCY_MCP_ENTRY)
        return {"status": "success", "report": report_content}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --------- 工单 Excel 上传 ---------
@app.post("/api/tickets/upload")
async def upload_tickets_excel(file: UploadFile = File(...)):
    """
    上传 Excel 文件解析工单数据

    Excel 格式要求：
    - 第一行为表头
    - 必须包含"主要内容"列
    - 可选列：序号、被反映街乡镇、所在村社区、二级承办单位简称

    Returns:
        解析后的工单数据列表
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx 或 .xls）")

    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel 文件为空或只有表头")

        headers = [str(h).strip() if h else "" for h in rows[0]]

        if "主要内容" not in headers:
            raise HTTPException(status_code=400, detail="Excel 必须包含主要内容列")

        tickets: List[Dict[str, Any]] = []
        for idx, row in enumerate(rows[1:], start=1):
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if header and col_idx < len(row):
                    value = row[col_idx]
                    row_dict[header] = str(value).strip() if value is not None else ""

            if not row_dict.get("主要内容"):
                continue

            ticket = {
                "序号": row_dict.get("序号", str(idx)),
                "主要内容": row_dict.get("主要内容", ""),
                "被反映街乡镇": row_dict.get("被反映街乡镇", ""),
                "所在村社区": row_dict.get("所在村社区", ""),
                "二级承办单位简称": row_dict.get("二级承办单位简称", ""),
                "建议处置部门": "",
                "派单理由": "",
                "历史工单": "",
                "规则依据": "",
                "备注": ""
            }
            tickets.append(ticket)

        wb.close()

        if not tickets:
            raise HTTPException(status_code=400, detail="未解析到有效工单数据")

        return {"status": "success", "count": len(tickets), "tickets": tickets}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel 解析失败: {str(e)}")


# --------- Dify API 代理端点 ---------
@app.post("/api/dify/upload")
async def dify_upload_proxy(
    app_type: str = Form(...),
    file: UploadFile = File(...),
    user: Optional[str] = Form(None)
):
    """
    Dify 文件上传代理

    Args:
        app_type: 应用类型 (order_recognition, element_extraction, dispatch_assistant, address_recognition)
        file: 上传的文件
        user: 用户标识（可选）

    Returns:
        Dify API 响应
    """
    try:
        result = await dify_proxy.upload_file(app_type, file, user)
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件上传失败: {str(e)}")


@app.post("/api/dify/workflow")
async def dify_workflow_proxy(request: DifyWorkflowRequest):
    """
    Dify 工作流运行代理（非流式）

    Args:
        request: 工作流请求参数

    Returns:
        Dify API 响应
    """
    try:
        result = await dify_proxy.run_workflow(
            app_type=request.app_type,
            inputs=request.inputs,
            response_mode=request.response_mode,
            user=request.user,
            conversation_id=request.conversation_id
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"工作流执行失败: {str(e)}")


@app.post("/api/dify/workflow/stream")
async def dify_workflow_stream_proxy(request: DifyWorkflowRequest):
    """
    Dify 工作流运行代理（流式）

    Args:
        request: 工作流请求参数

    Returns:
        流式响应
    """
    try:
        async def stream_generator():
            async for chunk in dify_proxy.run_workflow_stream(
                app_type=request.app_type,
                inputs=request.inputs,
                user=request.user,
                conversation_id=request.conversation_id
            ):
                yield chunk

        return StreamingResponse(
            stream_generator(),
            media_type="text/event-stream"
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"流式工作流执行失败: {str(e)}")

if __name__ == "__main__":
    uvicorn.run("web_server:app", host="0.0.0.0", port=8889, reload=True)
